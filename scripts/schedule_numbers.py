#!/usr/bin/env python3
"""Align the IRCB Schedule workbook to the feed and emit real episode numbers.

The site has no real episode numbers. `src/data/numbering.ts` numbers feed episodes 1..N by
air date, but the feed starts at episode 85 and carries 162 minisodes, interviews, bonuses
and annuals that never consumed an episode number. The result drifts: the episode titled
"400 Episodes of FOMO" is labelled EP. 435 and "500 Episodes and We've Finally Figured Out
Comic Books" is labelled EP. 543.

The workbook has the real numbers. It records *recording* dates and the feed records *air*
dates, and the titles carry no "Episode N |" prefix to join on (upstream strips it), so the
join is a monotone alignment on date: each feed episode takes the latest unclaimed sheet row
recorded no more than LAG_MAX days before it aired. Nearly all land at exactly 3 days, the
Sunday-record/Wednesday-air gap. The lag spread is the check that matters, and `main()` prints
it: a match drifting out toward the 14-day limit is the shape a wrong match has.

Episodes that match no sheet row keep NO number. They are the separately-numbered minisodes
and the untitled one-offs, and the sheet lists them with a blank `Ep` — inventing a number
for them is what the current code already does wrong.

The workbook is not public. Pull it from Drive as xlsx (its native export truncates and
silently drops rows) and pass the path:

    python scripts/schedule_numbers.py ~/Downloads/schedule.xlsx

Writes data/episode-numbers.csv, which is checked in and reviewable in a diff so the pipeline
never needs a live Sheets dependency.
"""
import csv
import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import xml.etree.ElementTree as ET

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CORE = ROOT / "public/d/core.json"
RSS_URL = "https://feeds.simplecast.com/U93zjuSN"
OUT = ROOT / "data/episode-numbers.csv"

# Recorded Sunday, aired Wednesday. 10 is the longest real gap observed (a holiday week);
# anything past 14 is a different episode and must not be claimed.
LAG_MAX = 14
TABS = (("Old Recording Dates", "Recording Date"), ("Schedule", "Rec. Date"))


def _ep(v):
    """(number, is_an_episode). A fractional Ep — 475.1, 522.1 — is a *skipped* week.

    They are labelled "NO EPISODE IN JULY?" in the 2025 run and left blank in the 2026 one,
    and they carry no host and no topic. Rounding them down to 475 or 522 hands the real
    episode a recording date belonging to a week it was not recorded in.
    """
    try:
        f = float(str(v).strip())
    except (TypeError, ValueError):
        return None, False
    return int(f), f == int(f)


def read_sheet(xlsx):
    """One row per week in order, ascending by episode number.

    `Rec. Date` is overwritten with "Done"/"DONE" once a recording is in the can, so the date
    is gone by the time the episode airs. Requiring one silently dropped 33 rows — which is
    every missing number on the site, exactly — and, worse, slid later numbers onto earlier
    episodes, because the skipped-week rows kept their dates and got claimed in their place.

    A missing date is recoverable because the rows are a weekly timeline and the skipped weeks
    are rows too: walk 7 days per row from the last dated row. Across the 129 dated rows either
    side of these gaps in the numbering era that depends on the sheet (Ep 400+, where the RSS
    title no longer states a number), the walk lands exactly on the next dated row every time
    but twice — one June 2024 stretch that shifted a day. The alignment window is 14 days, so a
    reconstruction has to be a fortnight wrong before it can pick the wrong episode.
    """
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    rows = {}
    for tab, datecol in TABS:
        ws = wb[tab]
        it = ws.iter_rows(values_only=True)
        hdr = [str(c).strip() if c is not None else "" for c in next(it)]
        week = None
        for raw in it:
            if all(c is None or str(c).strip() == "" for c in raw):
                continue
            r = dict(zip(hdr, raw))
            ep, is_episode = _ep(r.get("Ep"))
            if ep is None:
                continue
            rec = r.get(datecol)
            if isinstance(rec, datetime):
                week = rec
            elif week is None:
                continue                      # nothing to count from yet
            else:
                week = rec = week + timedelta(days=7)
            if not is_episode:
                continue                      # a skipped week holds its place and nothing else
            # Four numbers appear twice, rescheduled. The later row is the one that happened.
            if ep not in rows or rec > rows[ep]["rec"]:
                rows[ep] = {"ep": ep, "rec": rec, "topic": r.get("Topic")}
    return [rows[k] for k in sorted(rows)]


def align(sheet, feed):
    """Monotone alignment: numbers ascend with air date, and no row is claimed twice."""
    out, cursor = [], 0
    for pos, ep in enumerate(feed, start=1):
        air = datetime.fromisoformat(ep["date"][:19].replace("Z", ""))
        best = None
        for j in range(cursor, len(sheet)):
            lag = (air - sheet[j]["rec"]).days
            if lag < 0:
                break
            if lag <= LAG_MAX:
                best = (j, sheet[j], lag)
        if best:
            j, row, lag = best
            out.append({"ep": row["ep"], "key": ep["key"], "date": ep["date"][:10],
                        "title": ep["title"] or "", "shown_as": pos,
                        "rec_date": row["rec"].date().isoformat(), "lag_days": lag,
                        "topic": (row["topic"] or "").strip()})
            cursor = j + 1
    return out


def check(matched):
    """The alignment is only trustworthy if it is monotone and every lag is sane."""
    eps = [m["ep"] for m in matched]
    assert eps == sorted(eps), "episode numbers must ascend with air date"
    assert len(set(eps)) == len(eps), "an episode number was claimed twice"
    keys = [m["key"] for m in matched]
    assert len(set(keys)) == len(keys), "a feed episode was numbered twice"
    assert all(0 <= m["lag_days"] <= LAG_MAX for m in matched), "lag outside the sane window"


TITLE_NUM = re.compile(r"^\s*(?:i read comic books\s+)?episode\s+(\d+)\s*\|", re.I)


def feed_title_numbers():
    """The number the show itself printed in the RSS title, until it stopped in Jan 2024.

    This is the authoritative source wherever it exists, and it is also the only independent
    check on the sheet alignment — they agree on all 300 rows they share.
    """
    import urllib.request
    with urllib.request.urlopen(RSS_URL) as resp:
        tree = ET.fromstring(resp.read())
    out = {}
    for item in tree.iter("item"):
        title = item.findtext("title") or ""
        m = TITLE_NUM.match(title)
        if m:
            out[_titlekey(title)] = int(m.group(1))
    return out


def _titlekey(t):
    t = TITLE_NUM.sub("", t or "")
    return re.sub(r"[^a-z0-9]+", "", t.lower())


def main(xlsx):
    core = json.loads(CORE.read_text())
    feed = sorted([e for e in core["episodes"] if e.get("showId") and e.get("date")],
                  key=lambda e: e["date"])
    sheet = read_sheet(xlsx)
    matched = align(sheet, feed)
    check(matched)

    stated = feed_title_numbers()
    disagree = []
    for m in matched:
        n = stated.get(_titlekey(m["title"]))
        m["feed_title_ep"] = n if n is not None else ""
        if n is not None and n != m["ep"]:
            disagree.append((m["title"], m["ep"], n))
        if n is not None:
            m["ep"] = n                      # the show's own number always wins
            m["source"] = "feed-title"
        else:
            m["source"] = "schedule-sheet"

    OUT.parent.mkdir(exist_ok=True)
    cols = ["ep", "source", "shown_as", "delta", "date", "title", "feed_title_ep",
            "rec_date", "lag_days", "topic", "key"]
    with open(OUT, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for m in matched:
            w.writerow({**m, "delta": m["shown_as"] - m["ep"]})

    wrong = sum(1 for m in matched if m["shown_as"] != m["ep"])
    deltas = {m["shown_as"] - m["ep"] for m in matched}
    confirmed = sum(1 for m in matched if m["source"] == "feed-title")
    print(f"sheet rows usable: {len(sheet)}   feed episodes: {len(feed)}")
    print(f"numbered: {len(matched)}   unnumbered -> bonus: {len(feed) - len(matched)}")
    print(f"  confirmed by the show's own RSS title : {confirmed}")
    print(f"  from the schedule sheet alone         : {len(matched) - confirmed}  <- review these")
    print(f"  the two sources DISAGREE on           : {len(disagree)}")
    for t, a, b in disagree[:10]:
        print(f"      {t[:50]!r}: sheet {a}, feed title {b}")
    print(f"currently displayed WRONG: {wrong}   distinct offsets: {len(deltas)} "
          f"({min(deltas)}..{max(deltas)})")
    print(f"newest numbered episode: EP. {matched[-1]['ep']} (site shows {matched[-1]['shown_as']})")
    print(f"→ {OUT.relative_to(ROOT)}")


def selfcheck():
    """The two rules that recover a date, on a sheet small enough to read.

    `check()` guards the alignment on real data every run, but it cannot see this class of
    fault: dropping a row produces an alignment that is monotone, unique and inside the lag
    window — valid in every way it knows to test, and quietly missing 33 episodes.
    """
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = TABS[0][0]
    ws.append(["Ep", TABS[0][1], "Topic"])
    for row in [
        [100, datetime(2020, 1, 5), "dated"],
        [101, "DONE", "recorded, date overwritten"],     # -> 2020-01-12, one week on
        [102, datetime(2020, 1, 19), "dated again"],     # a real date always wins
        [103, datetime(2020, 1, 26), "dated"],
        ["103.1", datetime(2020, 2, 2), None],           # a skipped week, not episode 103
        [104, "Done", "after the skipped week"],         # -> 2020-02-09, counting past it
    ]:
        ws.append(row)
    wb.create_sheet(TABS[1][0]).append(["Ep", TABS[1][1], "Topic"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    got = {r["ep"]: r["rec"].date().isoformat() for r in read_sheet(buf)}

    assert got[101] == "2020-01-12", f"'DONE' should count a week on from 100, got {got[101]}"
    assert got[102] == "2020-01-19", "a stated date must beat the reconstruction"
    assert got[103] == "2020-01-26", f"103.1 is a skipped week, not 103's date, got {got[103]}"
    assert got[104] == "2020-02-09", f"the skipped week still costs a week, got {got[104]}"
    assert sorted(got) == [100, 101, 102, 103, 104], f"103.1 is not an episode: {sorted(got)}"
    print("selfcheck ok")


if __name__ == "__main__":
    if len(sys.argv) == 2 and sys.argv[1] == "--selfcheck":
        selfcheck()
    elif len(sys.argv) != 2:
        sys.exit(__doc__)
    else:
        main(sys.argv[1])
